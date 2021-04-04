from openpyxl import Workbook


def creating_file(name_of_file, name_of_sheet):
    wb = Workbook()
    ws = wb.active
    ws.title = "{}".format(name_of_sheet)
    names_of_columns = ['KM', 'Speed', 'LL-L D1', 'LL-R D1', 'AL-L D1', 'AL-R D1', 'LL-L D2', 'LL-R D2', 'AL-L D2',
                        'AL-R D2', 'LL-L D3', 'LL-R D3', 'AL-L D3', 'AL-R D3', 'Gage 1', 'Gage 2', 'Sup.', 'Twist/3m',
                        'Gage 1f', 'Gage 2f', 'Sup.f', 'Set 1 L', 'Set 1 R', 'Set 2 L', 'Set 2 R']
    for i in range(25):
        ws.cell(row=1, column=i+1, value=names_of_columns[i])

    wb.save(filename='{}.xlsx'.format(name_of_file))
    return name_of_file, name_of_sheet


def create_ecart_file(name_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet"
    name_columns = ["PK_début", "PK_fin", "LL-L D1", "LL-LH D1", "Diff LL-L", "LL-R D1", "LL-RH D1", "Diff LL-R",
                    "AL-L D1", "AL-LH D1", "Diff AL-L", "AL-R D1", "AL-RH D1", "Diff AL-R", "Twist", "Twist H",
                    "Diff twist"]
    for i in range(25):
        ws.cell(row=1, column=i+1, value=name_columns[i])
    wb.save(filename='{}.xlsx'.format(name_file))
    return name_file
