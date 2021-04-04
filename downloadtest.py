import excel_layout
import openpyxl
from tkinter import filedialog
from openpyxl import Workbook


def replace(m):
    m = m.replace("/", "//")
    return m


def transform(b):
    b = b * 10**-3
    return b


def lines_function(p):
    with open("{}".format(replace(p)), "r") as file:
        lines = file.readlines()[3:]
    return lines


def set_up_lines(line):
    line = line.split(";")
    for i in range(len(line)):
        line[i] = line[i].strip()
    j = 0
    while j <= 6:
        line.pop(0)
        j += 1
    line.pop(1)
    line[0] = str(float(line[0]) + transform(float(line[1])))
    line.pop(1)
    return line


def line_start_end(num_line_start, num_line_end, name_file, path_of_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "{}".format("Feuil1")
    names_of_columns = ['KM', 'Speed', 'LL-L D1', 'LL-R D1', 'AL-L D1', 'AL-R D1', 'LL-L D2', 'LL-R D2', 'AL-L D2',
                        'AL-R D2', 'LL-L D3', 'LL-R D3', 'AL-L D3', 'AL-R D3', 'Gage 1', 'Gage 2', 'Sup.', 'Twist/3m',
                        'Gage 1f', 'Gage 2f', 'Sup.f', 'Set 1 L', 'Set 1 R', 'Set 2 L', 'Set 2 R']
    for i in range(25):
        ws.cell(row=1, column=i+1, value=names_of_columns[i])

    a = 2
    if num_line_start > num_line_end:
        num_line_start, num_line_end = num_line_end, num_line_start
        for i in lines_function(path_of_file):
            if len(set_up_lines(i)) == 25 and num_line_start <= float(set_up_lines(i)[0]) <= num_line_end:
                for j in range(25):
                    ws.cell(row=a, column=j+1, value=set_up_lines(i)[j])
                a += 1
            if set_up_lines(i)[0] == num_line_start:
                break
        dir_name = filedialog.askdirectory()
        wb.save(filename='{}\\{}.xlsx'.format(dir_name, name_file))
    else:
        for i in lines_function(path_of_file):
            if len(set_up_lines(i)) == 25 and num_line_start <= float(set_up_lines(i)[0]) <= num_line_end:
                for j in range(25):
                    ws.cell(row=a, column=j+1, value=set_up_lines(i)[j])
                a += 1
            if set_up_lines(i)[0] == num_line_end:
                break
        dir_name = filedialog.askdirectory()
        wb.save(filename='{}\\{}.xlsx'.format(dir_name, name_file))


