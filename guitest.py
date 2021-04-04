import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
import numpy


def replace(m):
    m = m.replace("/", "//")
    return m


# def create_ecart_file(name_file):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "sheet1"
#     name_columns = ["PK_début", "PK_fin", "LL-L D1", "LL-LH D1", "Diff LL-L", "LL-R D1", "LL-RH D1", "Diff LL-R",
#                     "AL-L D1", "AL-LH D1", "Diff AL-L", "AL-R D1", "AL-RH D1", "Diff AL-R", "Twist", "Twist H",
#                     "Diff twist"]
#     for i in range(17):
#         ws.cell(row=1, column=i+1, value=name_columns[i])
#     wb.save(filename='{}.xlsx'.format(name_file))
#     return name_file


def ecart_type_function(start, path1, path2,devide):
    division = devide - 1
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    name_columns = ["PK_début", "PK_fin", "LL-L D1", "LL-LH D1", "Diff LL-L", "LL-R D1", "LL-RH D1", "Diff LL-R",
                    "AL-L D1", "AL-LH D1", "Diff AL-L", "AL-R D1", "AL-RH D1", "Diff AL-R", "Twist", "Twist H",
                    "Diff twist"]
    for i in range(17):
        ws.cell(row=1, column=i+1, value=name_columns[i])

    the_file = openpyxl.load_workbook("{}".format(replace(path1)))
    current_sheet = the_file["{}".format(the_file.sheetnames[0])]
    a = 0
    columns = 0
    row_iteration = 2
    maxrow = current_sheet.max_row
    # to know which column the line starts
    for row in current_sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        a = a+1
        for value in row:
            if value == start:
                columns = a

    while columns + division <= maxrow:
        list_of_dev = []
        j = 3
        start_value = current_sheet.cell(row=columns, column=1).value
        end_value = current_sheet.cell(row=columns+division, column=1).value
        p = 3
        while j <= 6:
            list_of_values = []
            for row in current_sheet.iter_rows(min_row=columns, max_row=columns+division, min_col=j, max_col=j):
                for value in row:
                    list_of_values.append(value.value)
            final_value = numpy.std(list_of_values)
            list_of_dev.append(final_value)
            list_value = []
            if j == 6:
                for row in current_sheet.iter_rows(min_row=columns, max_row=columns+division, min_col=18, max_col=18):
                    for value in row:
                        list_value.append(value.value)
                final_value = numpy.std(list_value)
                list_of_dev.append(final_value)
            j = j + 1

        ws.cell(row=row_iteration, column=1, value=start_value)
        ws.cell(row=row_iteration, column=2, value=end_value)

        for i in list_of_dev:
            ws.cell(row=row_iteration, column=p, value=i)
            p = p + 3
        columns = columns + devide
        row_iteration = row_iteration + 1

    the_file_2 = openpyxl.load_workbook("{}".format(replace(path2)))
    current_sheet_2 = the_file_2["{}".format(the_file_2.sheetnames[0])]
    qbt = 0
    columns_2 = 0
    row_iteration_2 = 2
    # to know which column the line starts
    for row_2 in current_sheet_2.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        qbt = qbt+1
        for value_2 in row_2:
            if value_2 == start:
                columns_2 = qbt

    while columns_2 - division >= 2:
        list_of_dev_2 = []
        azt = 3
        qsx = 4
        while azt <= 6:
            list_of_values_2 = []
            for row_2 in current_sheet_2.iter_rows(min_row=columns_2-division, max_row=columns_2, min_col=azt, max_col=azt):
                for value_2 in row_2:
                    list_of_values_2.append(value_2.value)
            final_value_2 = numpy.std(list_of_values_2)
            list_of_dev_2.append(final_value_2)
            list_value_2 = []
            if azt == 6:
                for row_2 in \
                        current_sheet_2.iter_rows(min_row=columns_2-division, max_row=columns_2, min_col=18, max_col=18):
                    for value_2 in row_2:
                        list_value_2.append(value_2.value)
                final_value_2 = numpy.std(list_value_2)
                list_of_dev_2.append(final_value_2)
            azt = azt + 1
        for i in list_of_dev_2:
            ws.cell(row=row_iteration_2, column=qsx, value=i)
            qsx = qsx + 3
        columns_2 = columns_2 - devide
        row_iteration_2 = row_iteration_2 + 1
    w = 3
    while w <= 17:
        m = 2
        for row_3 in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=w, max_col=w):
            array_1 = []
            for value_3 in row_3:
                array_1.append(value_3.value)
        w = w + 1
        for row_3 in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=w, max_col=w):
            array_2 = []
            for value_3 in row_3:
                array_2.append(value_3.value)
        w = w + 1
        for i in range(len(array_1)):
            ws.cell(row=m, column=w, value=abs(array_1[i]-array_2[i]))
            m = m + 1
        w = w + 1

    dir_name = filedialog.askdirectory()
    wb.save(filename='{}\\{}.xlsx'.format(dir_name, "Ecart_type"))





